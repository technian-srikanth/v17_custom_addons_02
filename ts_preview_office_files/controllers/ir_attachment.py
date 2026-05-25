from odoo import http
from odoo.http import request

import base64
import csv
import io
import json
import openpyxl


class AttachmentPreviewController(http.Controller):

    def _json_response(self, data, status=200):

        return request.make_response(

            json.dumps(data),

            headers=[
                ("Content-Type", "application/json")
            ],

            status=status
        )


    def _pdf_response(self, pdf_data):

        return request.make_response(

            pdf_data,

            headers=[

                ("Content-Type", "application/pdf"),

                ("Cache-Control", "public, max-age=3600"),
            ]
        )

    def _processing_response(self):

        return self._json_response({

            "status": "processing",

            "message":
                "Preview is still generating"

        }, status=202)


    def _get_preview_cache(self, attachment):

        cache_key = (
            f"{attachment.id}_{attachment.checksum}"
        )

        cached_name = (
            f"cache_{cache_key}.pdf"
        )

        return request.env[
            'ir.attachment'
        ].sudo().search([

            ('name', '=', cached_name)

        ], limit=1)

    @http.route(
        '/sheet/preview/<int:attachment_id>',
        auth='user',
        type='http'
    )
    def preview_attachment(self, attachment_id):

        attachment = request.env[
            'ir.attachment'
        ].sudo().browse(attachment_id)

        if (
            not attachment.exists()
            or not attachment.datas
        ):

            return self._json_response({
                "sheets": []
            })

        file_data = base64.b64decode(
            attachment.datas
        )

        filename = (
            attachment.name or ""
        ).lower()

        try:

            file_input = io.BytesIO(
                file_data
            )

            if (
                filename.endswith(".xlsx")
                or filename.endswith(".xlsm")
                or filename.endswith(".xlsb")
                or filename.endswith(".xls")
            ):

                wb = openpyxl.load_workbook(

                    file_input,

                    read_only=True,

                    data_only=True
                )

                sheets_data = []

                MAX_ROWS = 500

                for sheet_name in wb.sheetnames:

                    sheet = wb[sheet_name]

                    rows = []

                    for i, row in enumerate(

                        sheet.iter_rows(
                            values_only=True
                        )
                    ):

                        if i >= MAX_ROWS:
                            break

                        rows.append([

                            str(cell)
                            if cell is not None
                            else ""

                            for cell in row
                        ])

                    sheets_data.append({

                        "name": sheet_name,

                        "rows": rows
                    })

                return self._json_response({

                    "sheets": sheets_data
                })

            elif filename.endswith(".csv"):

                file_content = file_data.decode(

                    "utf-8",

                    errors="ignore"
                )

                csv_file = io.StringIO(
                    file_content
                )

                try:

                    dialect = (

                        csv.Sniffer().sniff(
                            file_content[:1024]
                        )

                        if file_content
                        else csv.excel
                    )

                    reader = csv.reader(
                        csv_file,
                        dialect
                    )

                except Exception:

                    csv_file.seek(0)

                    reader = csv.reader(
                        csv_file
                    )

                rows = list(reader)

                return self._json_response({

                    "sheets": [{

                        "name": attachment.name,

                        "rows": rows
                    }]
                })

            return self._json_response({

                "error":
                    "Unsupported file type"
            })

        except Exception as e:

            return self._json_response({

                "error": str(e)
            })


    @http.route(
        '/docx/preview/<int:attachment_id>',
        auth='user',
        type='http'
    )
    def docx_preview(self, attachment_id):

        attachment = request.env[
            'ir.attachment'
        ].sudo().browse(attachment_id)

        if not attachment.exists():

            return request.not_found()

        preview_cache = self._get_preview_cache(
            attachment
        )

        if not preview_cache:

            return self._processing_response()

        pdf_data = base64.b64decode(
            preview_cache.datas
        )

        return self._pdf_response(
            pdf_data
        )

    @http.route(
        '/ppt/preview/<int:attachment_id>',
        auth='user',
        type='http'
    )
    def ppt_preview(self, attachment_id):

        attachment = request.env[
            'ir.attachment'
        ].sudo().browse(attachment_id)

        if not attachment.exists():

            return request.not_found()

        preview_cache = self._get_preview_cache(
            attachment
        )

        if not preview_cache:

            return self._processing_response()

        pdf_data = base64.b64decode(
            preview_cache.datas
        )

        return self._pdf_response(
            pdf_data
        )